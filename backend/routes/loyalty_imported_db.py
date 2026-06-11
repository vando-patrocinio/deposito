"""Base de Dados de Clientes importada via XLSX (Atlaz/ERP externo).

Permite enriquecer o cadastro com:
  - Data de instalação real (Atlaz API não fornece)
  - Data de ativação / cadastro / cancelamento
  - Endereço completo, plano, status, telefones

Fluxo:
  1. Admin faz upload do .xlsx em POST /api/customer/loyalty-db/import
  2. Sistema salva cada linha em `loyalty_imported_db` (1 doc por CPF)
  3. Faz match por CPF (`document`) em `subscribers` e popula
     `installation_date` quando ausente
  4. Painel "Base de Dados" lista todas as linhas com busca/paginação
"""
from __future__ import annotations


NERVOUS_METADATA = {
    "owner": "platform-team",
    "domain": "infra",
    "criticality": "medium",
    "emits_events": False,
    "event_types": [],
    "company_id_required": True,
}

import io
import logging
import re
from datetime import datetime, timezone
from typing import Any, Optional

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from pymongo import UpdateOne

from core import DEMO_COMPANY_ID, require_role
from database import db

logger = logging.getLogger("ponto.loyalty_db")
router = APIRouter(prefix="/api/customer/loyalty-db", tags=["loyalty-imported-db"])


# Headers esperados (case-insensitive) → chave normalizada interna
HEADER_MAP = {
    "id": "external_id",
    "código de identificação": "code",
    "codigo de identificacao": "code",
    "nome": "name",
    "cpf/cnpj": "document",
    "rg/ie": "rg_ie",
    "plano": "plan_name",
    "mensalidade": "monthly_fee",
    "isento": "exempt",
    "login": "login",
    "senha": "password",
    "vencimento": "due_day",
    "status": "status",
    "data cadastro": "registration_date",
    "data de ativação": "activation_date",
    "data de ativacao": "activation_date",
    "data de cancelamento": "cancellation_date",
    "data agendada para instalação": "scheduled_install_date",
    "data agendada para instalacao": "scheduled_install_date",
    "data de instalação": "installation_date",
    "data de instalacao": "installation_date",
    "endereço": "street",
    "endereco": "street",
    "número": "number",
    "numero": "number",
    "complemento": "complement",
    "referência": "reference",
    "referencia": "reference",
    "bairro": "district",
    "cidade": "city",
    "estado": "state",
    "cep": "zip_code",
    "ssid": "ssid",
    "equipamento": "equipment",
    "comodato": "comodato",
    "vendedor": "seller",
    "instalador": "installer",
    "telefone 1": "phone1",
    "telefone 2": "phone2",
    "telefone 3": "phone3",
    # iter215s — novos campos do full.xlsx
    "emails": "email",
    "data de nascimento": "birth_date",
    "nome resumido": "name_short",
    "pai": "father_name",
    "mãe": "mother_name",
    "mae": "mother_name",
    "coordenadas": "coordinates",
    "chamados abertos": "tickets_open",
    "chamados fechados": "tickets_closed",
    "títulos pagos": "invoices_paid",
    "titulos pagos": "invoices_paid",
    "títulos à vencer": "invoices_due",
    "titulos a vencer": "invoices_due",
    "títulos vencidos": "invoices_overdue",
    "titulos vencidos": "invoices_overdue",
    "total vencido": "total_overdue",
}


def _digits(s: Any) -> str:
    return re.sub(r"\D+", "", str(s or ""))


def _normalize_phone_xlsx(s: Any) -> str:
    """Extrai dígitos e descarta valores espúrios (ex: '55' sozinho, que
    é só o código do país quando o XLSX vem com '+55 ()' vazio)."""
    d = _digits(s)
    if not d:
        return ""
    # Só "55" = código país sem número
    if d == "55":
        return ""
    # Menos de 10 dígitos = inválido (DDD + telefone tem mín 10)
    # Remove "55" do prefixo só pra checar tamanho real
    if d.startswith("55") and len(d) >= 12:
        return d
    if len(d) < 10:
        return ""
    return d


def _norm_cpf(s: Any) -> str:
    d = _digits(s)
    # CPF tem 11 dígitos, CNPJ 14. Mantém raw se for outro tamanho.
    if 1 <= len(d) <= 14:
        # Pad CPFs curtos a 11 dígitos (Excel come os zeros à esquerda)
        if len(d) <= 11:
            return d.zfill(11)
    return d


def _parse_dt(v: Any) -> Optional[str]:
    """Aceita datetime do Excel ou string 'DD/MM/YYYY[ HH:MM:SS]'."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        dt = v if v.tzinfo else v.replace(tzinfo=timezone.utc)
        return dt.isoformat()
    s = str(v).strip()
    # ISO já formatado?
    try:
        d = datetime.fromisoformat(s.replace("Z", "+00:00"))
        if d.tzinfo is None:
            d = d.replace(tzinfo=timezone.utc)
        return d.isoformat()
    except Exception:
        pass
    # DD/MM/YYYY [HH:MM:SS]
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y"):
        try:
            d = datetime.strptime(s, fmt).replace(tzinfo=timezone.utc)
            return d.isoformat()
        except ValueError:
            continue
    return None


@router.post("/import")
async def import_xlsx(
    file: UploadFile = File(...),
    user: dict = Depends(require_role("gestor")),
):
    """Importa arquivo XLSX da base de dados de clientes e popula
    `loyalty_imported_db`. Também enriquece `subscribers.installation_date`
    via match por CPF.
    """
    cid = user.get("company_id") or DEMO_COMPANY_ID
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Use arquivo .xlsx (Excel).")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content),
                                      read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Falha ao ler XLSX: {e}")
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        raise HTTPException(400, "Arquivo vazio.")

    # Mapeia índices das colunas reconhecidas
    col_idx: dict[str, int] = {}
    for i, h in enumerate(header):
        if h is None:
            continue
        key = HEADER_MAP.get(str(h).strip().lower())
        if key and key not in col_idx:
            col_idx[key] = i

    stats = {
        "rows_seen": 0,
        "rows_imported": 0,
        "rows_skipped_no_doc": 0,
        "subscribers_matched": 0,
        "subscribers_install_date_filled": 0,
    }

    now_iso = datetime.now(timezone.utc).isoformat()

    # iter215s — Em vez de truncar, fazemos upsert por external_id (preserva
    # campos vindos de outros arquivos como status, atividades, etc).
    cpf_to_dates: dict[str, dict] = {}
    upserted = 0
    pending_ops: list[UpdateOne] = []

    async def _flush(ops):
        if ops:
            await db.loyalty_imported_db.bulk_write(ops, ordered=False)

    for raw in rows_iter:
        stats["rows_seen"] += 1
        if not raw or all(v in (None, "") for v in raw):
            continue
        doc = {"company_id": cid, "imported_at": now_iso}
        for key, idx in col_idx.items():
            if idx >= len(raw):
                continue
            v = raw[idx]
            if v is None or v == "":
                continue
            if key == "document":
                doc[key] = _norm_cpf(v)
            elif key in ("registration_date", "activation_date",
                         "cancellation_date", "scheduled_install_date",
                         "installation_date"):
                d = _parse_dt(v)
                if d:
                    doc[key] = d
            elif key in ("monthly_fee", "total_overdue"):
                try:
                    doc[key] = float(v)
                except (ValueError, TypeError):
                    pass
            elif key in ("tickets_open", "tickets_closed",
                          "invoices_paid", "invoices_due",
                          "invoices_overdue"):
                try:
                    doc[key] = int(float(v))
                except (ValueError, TypeError):
                    pass
            elif key in ("due_day", "number"):
                try:
                    doc[key] = int(str(v).strip()) if str(v).strip() else None
                except ValueError:
                    doc[key] = str(v).strip()
            elif key.startswith("phone"):
                d = _normalize_phone_xlsx(v)
                if d:
                    doc[key] = d
            else:
                doc[key] = str(v).strip() if not isinstance(v, str) else v.strip()

        cpf = doc.get("document") or ""
        if not cpf or len(cpf) < 11:
            stats["rows_skipped_no_doc"] += 1
            continue

        # iter215s — upsert via bulk: chave preferencial é external_id (mais
        # único), fallback pra (company_id, document)
        ext_id = doc.get("external_id")
        if ext_id:
            flt = {"company_id": cid, "external_id": ext_id}
        else:
            flt = {"company_id": cid, "document": cpf}
        pending_ops.append(UpdateOne(flt, {"$set": doc}, upsert=True))
        upserted += 1
        if len(pending_ops) >= 500:
            await _flush(pending_ops)
            pending_ops = []
        stats["rows_imported"] = upserted

        # Pra enriquecimento: pega a melhor data (instalação > ativação > cadastro)
        best_date = (doc.get("installation_date")
                      or doc.get("activation_date")
                      or doc.get("registration_date"))
        if best_date:
            cpf_to_dates[cpf] = {
                "installation_date": best_date,
                "activation_date": doc.get("activation_date"),
                "registration_date": doc.get("registration_date"),
                "cancellation_date": doc.get("cancellation_date"),
                "status_imported": doc.get("status"),
            }

        # iter215s — bulk_docs não usado mais (upsert acima)

    # Flush final dos upserts pendentes
    await _flush(pending_ops)
    pending_ops = []
    stats["rows_imported"] = upserted

    # === Enriquece subscribers com installation_date via match por CPF ===
    if cpf_to_dates:
        # Pega subs que têm document em cpf_to_dates
        cpf_list = list(cpf_to_dates.keys())
        cursor = db.subscribers.find(
            {"company_id": cid, "document": {"$in": cpf_list}},
            {"_id": 0, "id": 1, "document": 1, "installation_date": 1},
        )
        async for s in cursor:
            stats["subscribers_matched"] += 1
            cpf = s["document"]
            data = cpf_to_dates.get(cpf) or {}
            update_set: dict = {}
            # Só preenche installation_date se estiver vazio
            if not s.get("installation_date") and data.get("installation_date"):
                update_set["installation_date"] = data["installation_date"]
                stats["subscribers_install_date_filled"] += 1
            if data.get("activation_date"):
                update_set["activation_date"] = data["activation_date"]
            if data.get("registration_date"):
                update_set["registration_date"] = data["registration_date"]
            if update_set:
                update_set["updated_at"] = now_iso
                await db.subscribers.update_one(
                    {"id": s["id"]}, {"$set": update_set},
                )

    # Salva metadados do último import
    await db.loyalty_import_log.insert_one({
        "company_id": cid,
        "filename": file.filename,
        "user": user.get("email") or user.get("id"),
        "at": now_iso,
        "stats": stats,
    })

    # iter215w — REGRA: toda importação invalida caches do painel Clientes
    # Fidelidade pra refletir os novos dados imediatamente em todas as abas.
    from routes.customer_loyalty import invalidate_loyalty_caches  # noqa
    invalidated = invalidate_loyalty_caches(cid)
    # Tb invalida análises IA antigas (regenera no próximo acesso)
    n_ai = await db.loyalty_ai_insights.delete_many({"company_id": cid})
    n_wb = await db.loyalty_winback_ready.delete_many({"company_id": cid})

    return {
        "ok": True,
        "filename": file.filename,
        "stats": stats,
        "cache_invalidated": {
            **invalidated,
            "ai_insights_cleared": n_ai.deleted_count,
            "winback_ready_cleared": n_wb.deleted_count,
        },
    }


@router.get("")
async def list_imported(
    q: Optional[str] = Query(None, description="Busca por nome/CPF/login"),
    status: Optional[str] = Query(None),
    city: Optional[str] = Query(None),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    user: dict = Depends(require_role("gestor")),
):
    """Lista paginada da base importada."""
    cid = user.get("company_id") or DEMO_COMPANY_ID
    flt: dict = {"company_id": cid}
    if status:
        flt["status"] = status
    if city:
        flt["city"] = city
    if q:
        rx = {"$regex": re.escape(q.strip()), "$options": "i"}
        flt["$or"] = [{"name": rx}, {"document": rx}, {"login": rx}]

    total = await db.loyalty_imported_db.count_documents(flt)
    cursor = db.loyalty_imported_db.find(
        flt, {"_id": 0}
    ).sort("name", 1).skip(skip).limit(limit)
    items = [doc async for doc in cursor]
    return {"total": total, "skip": skip, "limit": limit, "items": items}


@router.get("/stats")
async def imported_stats(user: dict = Depends(require_role("gestor"))):
    """KPIs agregados da base importada."""
    cid = user.get("company_id") or DEMO_COMPANY_ID
    total = await db.loyalty_imported_db.count_documents({"company_id": cid})
    by_status = await db.loyalty_imported_db.aggregate([
        {"$match": {"company_id": cid}},
        {"$group": {"_id": "$status", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
    ]).to_list(50)
    by_city = await db.loyalty_imported_db.aggregate([
        {"$match": {"company_id": cid}},
        {"$group": {"_id": "$city", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
        {"$limit": 20},
    ]).to_list(20)
    last_log = await db.loyalty_import_log.find_one(
        {"company_id": cid}, sort=[("at", -1)], projection={"_id": 0},
    )
    return {
        "total": total,
        "by_status": [{"status": s["_id"] or "—", "count": s["count"]}
                       for s in by_status],
        "by_city": [{"city": c["_id"] or "—", "count": c["count"]}
                     for c in by_city],
        "last_import": last_log,
    }
