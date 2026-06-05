"""Central de Compras — registro de compras de material com fluxo:

    COMPRA → ENTRADA NO ESTOQUE (Praça + responsável) → TÉCNICO → CLIENTE

Apenas:
- Administradores, gestores e super admin: veem TODAS as compras de TODAS
  as praças e podem confirmar entrada no estoque.
- Almoxarifes (collaborator com cargo=ALMOXARIFE + warehouse_praca_id):
  veem e lançam SOMENTE compras da própria praça.

Entrada do lançamento pode ser:
- Manual (form preenchido)
- Por upload de arquivo (PDF/XLS/DOC/JPG/PNG) → IA Claude extrai os dados.

Após confirmado, integra-se ao módulo `stok`:
- ONT: gera entrada em `stok_onts` (compatível com fluxo existente)
- Insumo: incrementa estoque do insumo na praça
- Outros: apenas registra (sem afetar `stok_*`)
"""
from __future__ import annotations

import base64
import json
import logging
import re
import uuid
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from pydantic import BaseModel, Field

import cargo as cargo_mod
from core import (
    DEMO_COMPANY_ID,
    get_current_user,
    is_super_admin,
    now_iso,
    require_role,
)
from database import db

logger = logging.getLogger("ponto.purchases")

router = APIRouter(prefix="/api/purchases", tags=["purchases"])


# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------
PURCHASE_TYPES = ("ont", "insumo", "equipamento", "outros")


class PurchaseItem(BaseModel):
    description: str
    quantity: float = 1
    unit: Optional[str] = None
    unit_price: Optional[float] = None
    type: Optional[str] = None  # ont|insumo|equipamento|outros (por item)
    macs: Optional[List[str]] = None  # se tipo=ont, MACs
    sns: Optional[List[str]] = None   # iter197 — SNs (prevalente sobre MAC)
    insumo_id: Optional[str] = None   # iter200 — bate com INSUMO_CATALOG


class PurchaseCreate(BaseModel):
    type: str = Field(..., description="ont|insumo|equipamento|outros")
    praca_id: str
    responsible_collaborator_id: str
    supplier_name: Optional[str] = None
    invoice_number: Optional[str] = None
    invoice_date: Optional[str] = None  # YYYY-MM-DD
    total_value: Optional[float] = None
    items: List[PurchaseItem] = []
    notes: Optional[str] = None
    file_url: Optional[str] = None
    file_name: Optional[str] = None


class PurchaseConfirmIn(BaseModel):
    purchase_id: str


# ---------------------------------------------------------------------------
# Permission helpers
# ---------------------------------------------------------------------------
def _is_admin_like(user: dict) -> bool:
    return is_super_admin(user) or user.get("role") in (
        "administrador", "gestor", "auditor",
    )


async def _resolve_user_praca(user: dict) -> Optional[str]:
    """Se o user logado é colaborador almoxarife, retorna a praça vinculada."""
    cid = user.get("company_id") or DEMO_COMPANY_ID
    if _is_admin_like(user):
        return None  # admin/gestor: sem restrição de praça
    coll = await db.collaborators.find_one(
        {"company_id": cid, "user_id": user.get("id")},
        {"_id": 0, "cargo": 1, "warehouse_praca_id": 1},
    )
    if not coll:
        return None
    if coll.get("cargo") != cargo_mod.ALMOXARIFE:
        return None
    return coll.get("warehouse_praca_id")


def _require_purchase_access(user: dict) -> None:
    """Lança 403 se o user não pode acessar a Central de Compras."""
    if _is_admin_like(user):
        return
    # Almoxarifes acessam (a verificação de praça é feita por query)
    if user.get("role") in ("financeiro",):
        return
    # Demais: bloqueia
    raise HTTPException(403, "Sem permissão para Central de Compras")


def _normalize_mac(s: str) -> str:
    s = re.sub(r"[^0-9A-Fa-f]", "", s or "").upper()
    return ":".join(s[i:i + 2] for i in range(0, len(s), 2)) if s else ""


# ---------------------------------------------------------------------------
# Routes — Listing
# ---------------------------------------------------------------------------
@router.get("")
async def list_purchases(
    user: dict = Depends(get_current_user),
    praca_id: Optional[str] = None,
    status: Optional[str] = None,
    type_: Optional[str] = None,
    limit: int = 100,
) -> Dict[str, Any]:
    _require_purchase_access(user)
    cid = user.get("company_id") or DEMO_COMPANY_ID
    q: Dict[str, Any] = {"company_id": cid}

    # Almoxarife só vê própria praça
    user_praca = await _resolve_user_praca(user)
    if user_praca:
        q["praca_id"] = user_praca
    elif praca_id:
        q["praca_id"] = praca_id
    if status:
        q["status"] = status
    if type_:
        q["type"] = type_

    docs = await db.purchases.find(q, {"_id": 0}) \
        .sort("created_at", -1).limit(min(limit, 500)).to_list(500)

    # Enrich com nomes de praça/responsável
    if docs:
        praca_ids = {d.get("praca_id") for d in docs if d.get("praca_id")}
        coll_ids = {d.get("responsible_collaborator_id") for d in docs
                     if d.get("responsible_collaborator_id")}
        # Resolve nomes de praça via `pracas` (fonte canônica do tenant).
        # Fallback em `fin_filiais` apenas para compras legadas cujo
        # `praca_id` foi gravado antes da migração do dropdown.
        pracas = {p["id"]: p["name"] async for p in db.pracas.find(
            {"id": {"$in": list(praca_ids)}}, {"_id": 0, "id": 1, "name": 1})}
        missing = [pid for pid in praca_ids if pid and pid not in pracas]
        if missing:
            async for p in db.fin_filiais.find(
                {"id": {"$in": missing}}, {"_id": 0, "id": 1, "name": 1}):
                pracas.setdefault(p["id"], p["name"])
        colls = {c["id"]: c["name"] async for c in db.collaborators.find(
            {"id": {"$in": list(coll_ids)}}, {"_id": 0, "id": 1, "name": 1})}
        for d in docs:
            d["praca_name"] = pracas.get(d.get("praca_id"), "—")
            d["responsible_name"] = colls.get(
                d.get("responsible_collaborator_id"), "—")
    return {
        "items": docs,
        "is_warehouse_keeper": user_praca is not None,
        "user_praca_id": user_praca,
    }


@router.get("/by-invoice")
async def list_by_invoice(user: dict = Depends(get_current_user),
                           limit: int = Query(60, ge=1, le=300)) -> Dict[str, Any]:
    """iter203 — Agrupa compras pelo par (supplier_name, invoice_number).

    Útil para conciliação fiscal: uma única NF pode ter virado N lançamentos
    (multi-tipo iter202). Esta rota devolve cada NF como um card único com:
      - lista de lançamentos (purchase IDs)
      - total consolidado (soma dos lançamentos)
      - breakdown por tipo (ont/insumo/ferramenta/equipamento/outros)
      - status global (mais "atrasado" entre os lançamentos)
    """
    _require_purchase_access(user)
    cid = user.get("company_id") or DEMO_COMPANY_ID
    pipeline = [
        {"$match": {"company_id": cid}},
        {"$sort": {"created_at": -1}},
        {"$group": {
            "_id": {
                "supplier": {"$ifNull": ["$supplier_name", "—"]},
                "invoice":  {"$ifNull": ["$invoice_number", "—"]},
            },
            "purchases": {"$push": {
                "id": "$id",
                "type": "$type",
                "status": "$status",
                "total_value": "$total_value",
                "invoice_date": "$invoice_date",
                "file_name": "$file_name",
                "praca_id": "$praca_id",
                "responsible_collaborator_id": "$responsible_collaborator_id",
                "items_count": {"$size": {"$ifNull": ["$items", []]}},
                "notes": "$notes",
                "created_at": "$created_at",
                "confirmed_at": "$confirmed_at",
            }},
            "total":      {"$sum": {"$ifNull": ["$total_value", 0]}},
            "count":      {"$sum": 1},
            "last_date":  {"$max": "$created_at"},
            "invoice_date": {"$max": "$invoice_date"},
        }},
        {"$sort": {"last_date": -1}},
        {"$limit": limit},
    ]
    rows = await db.purchases.aggregate(pipeline).to_list(limit)

    # Resolve nomes de praça/responsável uma vez (cache)
    pracas = {p["id"]: p["name"] for p in await db.pracas.find(
        {"company_id": cid}, {"_id": 0, "id": 1, "name": 1}).to_list(500)}
    colls = {c["id"]: c["name"] for c in await db.collaborators.find(
        {"company_id": cid}, {"_id": 0, "id": 1, "name": 1}).to_list(500)}

    items: List[Dict[str, Any]] = []
    for r in rows:
        types_summary: Dict[str, int] = {}
        statuses: set[str] = set()
        for p in r["purchases"]:
            t = p.get("type") or "outros"
            types_summary[t] = types_summary.get(t, 0) + (p.get("items_count") or 0)
            statuses.add(p.get("status") or "")
            # enriquece com nomes
            p["praca_name"] = pracas.get(p.get("praca_id"))
            p["responsible_name"] = colls.get(p.get("responsible_collaborator_id"))
        # status global: pending > received > confirmed (worst-first)
        order = ["pending", "received", "confirmed"]
        global_status = next((s for s in order if s in statuses), "confirmed")
        items.append({
            "supplier_name": r["_id"]["supplier"],
            "invoice_number": r["_id"]["invoice"],
            "invoice_date": r.get("invoice_date"),
            "total_value": round(r["total"] or 0, 2),
            "count": r["count"],
            "purchases": r["purchases"],
            "types_summary": types_summary,
            "global_status": global_status,
            "last_date": r["last_date"],
        })
    return {"items": items, "total": len(items)}


@router.get("/refs")
async def get_refs(user: dict = Depends(get_current_user)) -> Dict[str, Any]:
    """Retorna praças, responsáveis (almoxarifes), fornecedores existentes
    e tipos para autocomplete no form.

    Praças vêm de `pracas` (sidebar → Cadastro → Praças). Esta é a fonte
    canônica do tenant; o dropdown "Praça destino" reflete exatamente o
    que está cadastrado na aba Praças.
    """
    _require_purchase_access(user)
    cid = user.get("company_id") or DEMO_COMPANY_ID
    pracas = await db.pracas.find(
        {"company_id": cid},
        {"_id": 0, "id": 1, "name": 1, "city": 1, "state": 1},
    ).sort("name", 1).to_list(500)
    colls = await db.collaborators.find(
        {"company_id": cid, "active": {"$ne": False}},
        {"_id": 0, "id": 1, "name": 1, "cargo": 1,
          "warehouse_praca_id": 1},
    ).sort("name", 1).to_list(500)
    suppliers = await db.fin_suppliers.find(
        {"company_id": cid, "active": {"$ne": False}},
        {"_id": 0, "id": 1, "name": 1, "document": 1, "cnpj": 1},
    ).sort("name", 1).to_list(500)
    return {"pracas": pracas, "collaborators": colls,
             "suppliers": suppliers,
             "types": list(PURCHASE_TYPES)}


def _norm_supplier(name: str) -> str:
    """Normaliza nome de fornecedor para busca fuzzy."""
    import unicodedata
    s = (name or "").strip().upper()
    s = "".join(c for c in unicodedata.normalize("NFD", s)
                  if unicodedata.category(c) != "Mn")
    s = re.sub(r"\b(LTDA|ME|EPP|S/A|SA|EIRELI|MEI)\b", "", s)
    s = re.sub(r"[^A-Z0-9]+", " ", s).strip()
    return s


async def _find_or_create_supplier(
    cid: str, name: str, user_email: str,
) -> Optional[str]:
    """Encontra fornecedor por nome normalizado; se não existir, cria.

    Retorna o `id` do fornecedor (ou None se nome vazio).
    """
    name = (name or "").strip()
    if not name:
        return None
    norm = _norm_supplier(name)
    if not norm:
        return None
    # Busca exata primeiro
    existing = await db.fin_suppliers.find_one(
        {"company_id": cid, "name": name}, {"_id": 0, "id": 1})
    if existing:
        return existing["id"]
    # Busca normalizada (compara em memória — fin_suppliers é pequena)
    async for s in db.fin_suppliers.find(
            {"company_id": cid}, {"_id": 0, "id": 1, "name": 1}):
        if _norm_supplier(s.get("name", "")) == norm:
            return s["id"]
    # Cria
    new_id = f"fsup-{uuid.uuid4().hex[:10]}"
    await db.fin_suppliers.insert_one({
        "id": new_id,
        "company_id": cid,
        "name": name[:200],
        "active": True,
        "created_at": now_iso(),
        "created_by": user_email,
        "auto_created_from": "central_compras",
    })
    logger.info("[purchases] fornecedor auto-criado: %s (%s) cid=%s",
                  name, new_id, cid)
    return new_id


# ---------------------------------------------------------------------------
# Routes — Manual create
# ---------------------------------------------------------------------------
@router.post("")
async def create_purchase(
    payload: PurchaseCreate,
    user: dict = Depends(get_current_user),
) -> Dict[str, Any]:
    _require_purchase_access(user)
    cid = user.get("company_id") or DEMO_COMPANY_ID
    if payload.type not in PURCHASE_TYPES:
        raise HTTPException(400, f"Tipo inválido. Use: {PURCHASE_TYPES}")

    # Almoxarife: força praça dele
    user_praca = await _resolve_user_praca(user)
    if user_praca and payload.praca_id != user_praca:
        raise HTTPException(403, "Você só pode lançar compras da sua praça")

    pid = f"pur-{uuid.uuid4().hex[:10]}"
    # Resolve/cria fornecedor automaticamente (link com fin_suppliers)
    supplier_id = await _find_or_create_supplier(
        cid, payload.supplier_name or "", user.get("email") or "?",
    )
    doc = {
        "id": pid,
        "company_id": cid,
        "type": payload.type,
        "praca_id": payload.praca_id,
        "responsible_collaborator_id": payload.responsible_collaborator_id,
        "supplier_name": payload.supplier_name,
        "supplier_id": supplier_id,
        "invoice_number": payload.invoice_number,
        "invoice_date": payload.invoice_date,
        "total_value": payload.total_value,
        "items": [it.dict() for it in payload.items],
        "notes": payload.notes,
        "file_url": payload.file_url,
        "file_name": payload.file_name,
        "status": "pending",  # pending -> confirmed
        "created_at": now_iso(),
        "created_by": user.get("email"),
        "created_by_id": user.get("id"),
    }
    await db.purchases.insert_one(dict(doc))
    doc.pop("_id", None)
    return doc


# ---------------------------------------------------------------------------
# Routes — Upload + AI extraction
# ---------------------------------------------------------------------------
@router.post("/upload-extract")
async def upload_extract(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user),
) -> Dict[str, Any]:
    """Upload de NF/recibo (PDF/imagem). IA Claude extrai os dados.

    Retorna um DRAFT (não persiste). Usuário revisa e confirma via POST /.
    """
    _require_purchase_access(user)
    raw = await file.read()
    if len(raw) > 10_000_000:
        raise HTTPException(413, "Arquivo > 10 MB")
    fname = (file.filename or "compra").lower()
    is_pdf = fname.endswith(".pdf") or raw[:5] == b"%PDF-"
    is_img = any(fname.endswith(e) for e in
                  (".jpg", ".jpeg", ".png", ".webp"))
    is_doc = any(fname.endswith(e) for e in
                  (".doc", ".docx", ".xls", ".xlsx"))
    if not (is_pdf or is_img or is_doc):
        raise HTTPException(415, "Tipo não suportado. Envie PDF, "
                                   "imagem ou planilha.")

    cid = user.get("company_id") or DEMO_COMPANY_ID
    # Extrair texto/imagem
    text_content = ""
    image_b64: Optional[str] = None
    try:
        if is_pdf:
            import pdfplumber
            import io
            with pdfplumber.open(io.BytesIO(raw)) as pdf:
                for page in pdf.pages[:5]:  # primeiros 5
                    t = page.extract_text() or ""
                    text_content += t + "\n"
        elif is_doc:
            # docx/xlsx: leitura simples
            try:
                if fname.endswith((".xls", ".xlsx")):
                    from openpyxl import load_workbook
                    import io
                    wb = load_workbook(io.BytesIO(raw), data_only=True)
                    for ws in wb.worksheets[:3]:
                        for row in ws.iter_rows(values_only=True,
                                                  max_row=200):
                            text_content += " | ".join(
                                str(c) for c in row if c) + "\n"
                else:
                    from docx import Document
                    import io
                    d = Document(io.BytesIO(raw))
                    for p in d.paragraphs:
                        text_content += p.text + "\n"
            except Exception:
                pass
        elif is_img:
            image_b64 = base64.b64encode(raw).decode()
    except Exception as e:
        logger.warning("[purchases] extract falhou: %s", e)

    # iter211q — Pré-OCR Vision para IMAGENS (1ª camada).
    # Antes da IA estruturar o JSON, peça SÓ a transcrição bruta do texto
    # da imagem. Vai alimentar:
    #   (a) Prompt principal (mais contexto)
    #   (b) text_content para o regex final
    #   (c) raw_text_preview salvo na compra (para reprocess futuro)
    if is_img and image_b64 and not text_content:
        try:
            from emergentintegrations.llm.chat import (
                ImageContent as _Img, LlmChat as _Llm,
                UserMessage as _Msg,
            )
            from services.ai_keys import resolve_keys as _resolve
            _keys = await _resolve(cid)
            _ak = (_keys.get("anthropic") or _keys.get("openai")
                       or _keys.get("gemini"))
            if _ak:
                ocr_prompt = (
                    "Transcreva TODO o texto visível desta nota fiscal/"
                    "recibo, linha por linha, preservando a ordem "
                    "vertical. NÃO interprete nem traduza. Inclua TODOS "
                    "os números de série (Nº Série/SN/S/N) na íntegra, "
                    "mesmo se forem dezenas. NÃO use markdown. NÃO use "
                    "JSON. Apenas o TEXTO BRUTO transcrito."
                )
                _ocr_chat = _Llm(
                    api_key=_ak,
                    session_id=f"ocr-{cid}-{uuid.uuid4().hex[:6]}",
                    system_message="Você transcreve texto de imagens fielmente.",
                ).with_model("anthropic", "claude-sonnet-4-5")
                ocr_resp = await _ocr_chat.send_message(_Msg(
                    text=ocr_prompt,
                    file_contents=[_Img(image_base64=image_b64)],
                ))
                text_content = (ocr_resp or "").strip()
                logger.info(
                    "[purchases iter211q] OCR Vision: %d chars",
                    len(text_content))
        except Exception as e:  # noqa: BLE001
            logger.debug("[purchases iter211q] pré-OCR falhou: %s", e)

    # Chama Claude Sonnet 4.5 para estruturar
    try:
        from emergentintegrations.llm.chat import (
            ImageContent, LlmChat, UserMessage,
        )
        from services.ai_keys import resolve_keys
        keys = await resolve_keys(cid)
        ai_key = (keys.get("anthropic") or keys.get("openai")
                       or keys.get("gemini"))
        if not ai_key:
            raise HTTPException(503, "IA não configurada — preencha manual")
        prompt = (
            "Extraia os dados desta nota fiscal / recibo / comprovante "
            "de compra de material para um provedor de internet "
            "brasileiro. Responda APENAS JSON válido (sem markdown):\n"
            "{\n"
            "  \"supplier_name\": \"...\",  // razão social/nome\n"
            "  \"invoice_number\": \"...\",\n"
            "  \"invoice_date\": \"YYYY-MM-DD\",\n"
            "  \"total_value\": 1234.56,\n"
            "  \"type\": \"ont\"|\"insumo\"|\"equipamento\"|\"ferramenta\"|\"outros\",\n"
            "    // TIPO DOMINANTE da nota (o que mais aparece)\n"
            "  \"items\": [{\n"
            "      \"description\": \"...\",\n"
            "      \"quantity\": 1,\n"
            "      \"unit\": \"un|m|cx|pç\",\n"
            "      \"unit_price\": 0.0,\n"
            "      \"type\": \"ont\"|\"insumo\"|\"equipamento\"|\"ferramenta\"|\"outros\",\n"
            "        // CLASSIFIQUE CADA ITEM:\n"
            "        // - ferramenta: alicate, chave, OTDR, decapador, máquina\n"
            "        //   de fusão, escada, multímetro, power meter etc.\n"
            "        // - insumo: splitter, conector, cordão óptico, drop,\n"
            "        //   cabo de rede, esticador, fibra (06FO/12FO/24FO),\n"
            "        //   adaptador, CTO sem splitter.\n"
            "        // - ont: ONU/ONT cliente (Huawei HG6145D, ZTE F660,\n"
            "        //   Fiberhome, Intelbras, Nokia G-1425G, FIBERHOME AC1200,\n"
            "        //   qualquer item com 'ONT', 'ONU', 'GPON', 'AC1200' OU\n"
            "        //   código de modelo (HG6145D, F660, F670L, G-1425G).\n"
            "        // - equipamento: OLT, switch, rack, nobreak, servidor,\n"
            "        //   GBIC/SFP, módulo BIDI, transceiver.\n"
            "      \"insumo_id\": null,\n"
            "        // SE for INSUMO, mapeie ao catálogo do sistema usando\n"
            "        // EXATAMENTE um destes IDs (ou null se não bater):\n"
            "        // 'drop' (cabo óptico drop FTTH, unit=m),\n"
            "        // 'cabo_rede' (cabo de rede UTP/Cat5e/Cat6, unit=m),\n"
            "        // 'conector_fast' (conector fast SC/APC ou SC/UPC),\n"
            "        // 'conector_fibra' (conector de fibra qualquer),\n"
            "        // 'esticador' (alça pré-formada/esticador de drop),\n"
            "        // 'conector_rede' (conector RJ45),\n"
            "        // 'fibra_06fo', 'fibra_12fo', 'fibra_24fo'\n"
            "        //   (cabo óptico backbone de 6/12/24 fibras).\n"
            "      \"sns\": [\"...\"],  // se ONT/EQUIPAMENTO, TODOS os SNs visíveis\n"
            "      \"macs\": [\"...\"]  // se ONT, MACs identificados (opcional)\n"
            "  }],\n"
            "  \"confidence\": 0.0-1.0,\n"
            "  \"reason\": \"breve explicação PT-BR do que extraiu\"\n"
            "}\n\n"
            "REGRAS IMPORTANTES:\n"
            "1. Se houver itens de tipos diferentes, escolha o TYPE dominante.\n"
            "2. NUNCA invente SN/MAC — só liste se estiver explícito.\n"
            "3. Para insumos, SEMPRE tente mapear ao catálogo (insumo_id).\n"
            "4. Quantidade deve ser numérica (1, 100, 305, 1000, etc.).\n"
            "5. unit_price = preço UNITÁRIO (não o total da linha).\n"
            "6. **CRÍTICO**: liste EM `items[]` TODAS as linhas da NF "
            "individualmente — NUNCA consolide ou agrupe itens. Se a NF "
            "tem 5 linhas (ex: ONT + cabo drop + conectores + alicate + "
            "splitter), o array `items` deve ter EXATAMENTE 5 entradas. "
            "O sistema cria automaticamente 1 lançamento por TIPO; quanto "
            "mais detalhado, melhor o estoque fica.\n"
            "7. **MUITO IMPORTANTE — SNs MÚLTIPLOS NO MESMO ITEM**: É comum "
            "1 LINHA da NF ter quantidade > 1 e vários SNs listados em "
            "sequência (vertical ou separados por vírgula). Exemplo real:\n"
            "   'FIBERHOME ONT AC1200 GPON HG6145D\n"
            "    Nº Série: FHTTC250CE0B, FHTTC250CE0C, FHTTC250CE14,\n"
            "    FHTTC250D38F, FHTTC250D394, FHTTC250D499, FHTTC250D4F7,\n"
            "    FHTTC250DAA5, FHTTC250DED9, FHTTC250DEF5    Qtd: 10'\n"
            "   → ISSO É **1 ITEM** com quantity=10 e sns=[10 SNs]. NÃO "
            "agrupe nem ignore os SNs! Cada SN abaixo de 'Nº Série:' é "
            "uma ONT física do mesmo modelo.\n"
            "8. Padrões de SN comuns: 'FHTT...', 'HUAW...', 'ZTEG...', "
            "'INVR...', 'ASTT...', 'ITBS...' (12-20 chars alfanuméricos). "
            "Capture TODOS no array `sns`.\n"
            "9. Se quantity=10 mas você só consegue ler 7 SNs nítidos por "
            "borrão da foto, retorne os 7 que conseguiu — NUNCA invente "
            "os 3 restantes. O sistema avisa o usuário.\n\n"
        )
        if text_content:
            prompt += f"TEXTO EXTRAÍDO:\n{text_content[:8000]}"
        chat = LlmChat(
            api_key=ai_key,
            session_id=f"purchase-{cid}-{uuid.uuid4().hex[:6]}",
            system_message="Você responde APENAS em JSON válido.",
        ).with_model("anthropic", "claude-sonnet-4-5")
        msg_args: Dict[str, Any] = {"text": prompt}
        # iter211q — Quando temos texto OCR pré-extraído (camada 1),
        # passa junto no prompt como referência — a IA tende a errar
        # menos os SNs longos quando vê a transcrição literal.
        if text_content:
            msg_args["text"] = (
                prompt
                + f"\n\nTEXTO BRUTO TRANSCRITO DA NF (para referência — "
                f"use isso para extrair SNs):\n```\n{text_content[:4000]}\n```"
            )
        if image_b64:
            msg_args["file_contents"] = [ImageContent(image_base64=image_b64)]
        resp = await chat.send_message(UserMessage(**msg_args))
        text = (resp or "").strip()
        if text.startswith("```"):
            text = re.sub(r"^```(?:json)?\s*|\s*```$", "", text,
                              flags=re.MULTILINE)
        draft = json.loads(text)
    except json.JSONDecodeError as e:
        logger.warning("[purchases] IA JSON inválido: %s", e)
        draft = {"confidence": 0.0,
                  "reason": "IA não conseguiu estruturar — preencha manual."}
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("[purchases] upload-extract falhou: %s", e)
        draft = {"confidence": 0.0, "reason": f"erro: {e!s}"}

    # iter211n — Reconciliação de SNs: se a IA disse "ONT qty=10 sns=[3]",
    # extrai SNs no texto bruto via regex e enriquece o item. Padrões comuns
    # de SN em NF (Fiberhome FHTTC..., Huawei HUAW..., ZTE ZTEG..., etc.).
    # iter211p — Para IMAGENS, faz 2ª chamada Vision focada SÓ em SNs quando
    # algum item ONT está com `quantity > len(sns)`.
    needs_sn_pass = False
    if isinstance(draft, dict):
        for it in (draft.get("items") or []):
            if (it.get("type") or "").lower() == "ont":
                qty = int(it.get("quantity") or 0)
                if qty > len(it.get("sns") or []):
                    needs_sn_pass = True
                    break

    if needs_sn_pass and image_b64:
        try:
            sn_prompt = (
                "Olhe esta NOTA FISCAL e extraia APENAS os números de série "
                "(SN) das ONTs/ONUs/equipamentos. Os SNs costumam:\n"
                "- Aparecer abaixo do nome do produto (Nº Série / SN /  S/N)\n"
                "- Listados verticalmente, um por linha, separados por "
                "vírgula ou ponto-e-vírgula\n"
                "- Padrões comuns: FHTT*, HUAW*, ZTEG*, INVR*, ASTT*, "
                "ITBS*, GPON*, SHTL*, FBHM*, INTB*\n"
                "- 10-20 caracteres alfanuméricos MAIÚSCULOS\n\n"
                "Responda APENAS JSON: {\"sns\": [\"FHTTC250CE0B\", ...]}\n"
                "Liste TODOS os SNs visíveis na imagem, mesmo que parciais."
                " NÃO invente.")
            sn_chat = LlmChat(
                api_key=ai_key,
                session_id=f"purchase-snpass-{cid}-{uuid.uuid4().hex[:6]}",
                system_message="Você responde APENAS em JSON válido.",
            ).with_model("anthropic", "claude-sonnet-4-5")
            sn_resp = await sn_chat.send_message(UserMessage(
                text=sn_prompt,
                file_contents=[ImageContent(image_base64=image_b64)],
            ))
            sn_text = (sn_resp or "").strip()
            if sn_text.startswith("```"):
                sn_text = re.sub(r"^```(?:json)?\s*|\s*```$", "", sn_text,
                                   flags=re.MULTILINE)
            sn_obj = json.loads(sn_text)
            extra = [str(s).strip().upper()
                       for s in (sn_obj.get("sns") or []) if s]
            if extra:
                # Anexa ao primeiro item ONT incompleto
                for it in (draft.get("items") or []):
                    if (it.get("type") or "").lower() != "ont":
                        continue
                    qty = int(it.get("quantity") or 0)
                    cur = [s.upper() for s in (it.get("sns") or []) if s]
                    if qty <= len(cur):
                        continue
                    seen = set(cur)
                    for sn in extra:
                        if sn not in seen:
                            cur.append(sn)
                            seen.add(sn)
                    it["sns"] = cur
                    logger.info(
                        "[purchases iter211p] 2ª pass Vision capturou %s SNs "
                        "(item '%s')", len(extra),
                        (it.get("description") or "")[:40])
                    break
                # Também alimenta text_content pra o regex pegar embaixo
                text_content = (text_content or "") + "\nSN_PASS: " \
                    + " ".join(extra)
        except Exception as e:  # noqa: BLE001
            logger.debug("[purchases iter211p] 2ª pass SN falhou: %s", e)

    if isinstance(draft, dict) and text_content:
        try:
            import re as _re
            # Pega "blocos" de SN consecutivos (>= 6 chars de A-Z+0-9, sem
            # espaços). Filtra os que parecem MAC ou data.
            SN_RE = _re.compile(r"\b[A-Z]{2,6}[A-Z0-9]{6,20}\b")
            STOPWORDS = {"CNPJ", "RUA", "AVENIDA", "CEP", "RECIBO",
                          "ORDEM", "FORNECEDOR", "CLIENTE", "TELEFONE",
                          "ITAU", "GPON", "WIFI", "AC1200", "BIDI"}
            # Lista de SNs candidatos no texto inteiro
            all_candidates = []
            for m in SN_RE.finditer((text_content or "").upper()):
                tok = m.group(0)
                if tok in STOPWORDS:
                    continue
                # Descarta tokens "modelo" típicos (HG6145D, F670L etc).
                if len(tok) <= 8:
                    continue
                all_candidates.append(tok)
            # Para cada item ONT com qty > len(sns), tenta enriquecer
            for it in (draft.get("items") or []):
                if (it.get("type") or "").lower() != "ont":
                    continue
                qty = int(it.get("quantity") or 0)
                cur_sns = [s for s in (it.get("sns") or []) if s]
                missing = max(0, qty - len(cur_sns))
                if missing <= 0:
                    continue
                # Sugere candidates que ainda não estão no item
                seen = {s.upper() for s in cur_sns}
                for cand in all_candidates:
                    if cand in seen:
                        continue
                    cur_sns.append(cand)
                    seen.add(cand)
                    if len(cur_sns) >= qty:
                        break
                it["sns"] = cur_sns
                if missing > 0 and len(it["sns"]) > len(seen) - missing:
                    logger.info(
                        "[purchases iter211n] regex enriqueceu SNs do item "
                        "'%s' — qty=%s sns=%s",
                        (it.get("description") or "")[:40], qty, len(it["sns"]))
        except Exception as e:  # noqa: BLE001
            logger.debug("[purchases iter211n] regex enrich falhou: %s", e)

    # iter202 — Multi-lançamento: agrupa items por TYPE em drafts separados
    # (uma única NF pode virar 2-3 lançamentos: ONTs + insumos + ferramentas
    # cada um indo pra sua coleção correta no estoque).
    drafts = _split_draft_by_type(draft) if isinstance(draft, dict) else [draft]

    return {
        "ok": True,
        "file_name": file.filename,
        "draft": drafts[0] if drafts else draft,  # compat: 1º draft no campo antigo
        "drafts": drafts,                          # iter202 — todos os drafts
        "draft_count": len(drafts),
        "raw_text_preview": text_content[:3000] if text_content else None,
    }


def _split_draft_by_type(draft: Dict[str, Any]) -> List[Dict[str, Any]]:
    """iter202 — Quebra 1 draft de extração da IA em N drafts (um por tipo).

    Estratégia:
      - Itens sem `type` herdam o `type` dominante do draft.
      - Items são agrupados por type → cada grupo vira um draft com:
        - mesmos metadados (fornecedor, NF, data, file_name)
        - `type` = type do grupo
        - `items` = só os items daquele tipo
        - `total_value` proporcional (soma dos unit_price × quantity do grupo).
      - Order: ont → insumo → ferramenta → equipamento → outros.
    """
    items = draft.get("items") or []
    if not items:
        return [draft]
    dominant_type = draft.get("type") or "outros"
    # 1) Agrupa por type efetivo
    groups: Dict[str, List[Dict[str, Any]]] = {}
    for it in items:
        t = (it.get("type") or dominant_type or "outros").lower()
        groups.setdefault(t, []).append(it)
    # 2) Se só há 1 grupo, devolve draft original (sem ruído de split)
    if len(groups) <= 1:
        return [draft]
    # 3) Ordem fixa (ont primeiro pois é o mais "frágil" no fluxo)
    order = ["ont", "insumo", "equipamento", "ferramenta", "outros"]
    sorted_types = sorted(groups.keys(),
                          key=lambda x: order.index(x) if x in order else 99)
    # 4) Constrói N drafts
    base_meta = {
        k: draft.get(k) for k in
        ("supplier_name", "invoice_number", "invoice_date",
         "confidence", "reason")
    }
    result: List[Dict[str, Any]] = []
    for t in sorted_types:
        group_items = groups[t]
        group_total = sum(
            float(i.get("unit_price") or 0) * float(i.get("quantity") or 1)
            for i in group_items
        )
        result.append({
            **base_meta,
            "type": t,
            "items": group_items,
            "total_value": round(group_total, 2) if group_total else None,
            "split_from_invoice": True,
            "split_part": f"{len(result) + 1}/{len(groups)}",
        })
    return result


# ---------------------------------------------------------------------------
# Routes — Confirm (integra com stok)
# ---------------------------------------------------------------------------
@router.post("/{purchase_id}/confirm")
async def confirm_purchase(
    purchase_id: str,
    user: dict = Depends(require_role("gestor")),
) -> Dict[str, Any]:
    """Confirma a compra e gera entradas em stok_onts/stok_stock conforme tipo.

    Apenas gestor/administrador pode confirmar (regra de negócio do estoque).
    """
    cid = user.get("company_id") or DEMO_COMPANY_ID
    p = await db.purchases.find_one({"id": purchase_id, "company_id": cid},
                                          {"_id": 0})
    if not p:
        raise HTTPException(404, "Compra não encontrada")
    if p.get("status") == "confirmed":
        raise HTTPException(400, "Compra já confirmada")

    items_imported = 0
    macs_imported: List[str] = []
    notes: List[str] = []

    if p["type"] == "ont":
        # iter197 — SN é o identificador prevalente. Aceita SN sozinho (gera
        # MAC placeholder "SN-..."), MAC sozinho, ou ambos pareados por índice.
        # Coleta pares (mac, sn) deduplicados por SN preferencialmente, MAC senão.
        seen_sns: set[str] = set()
        seen_macs: set[str] = set()
        pairs: List[Dict[str, Optional[str]]] = []  # [{mac, sn}]
        for it in (p.get("items") or []):
            macs_list = list(it.get("macs") or [])
            sns_list = list(it.get("sns") or [])
            n = max(len(macs_list), len(sns_list))
            for i in range(n):
                mac_raw = macs_list[i] if i < len(macs_list) else None
                sn_raw = sns_list[i] if i < len(sns_list) else None
                mac = _normalize_mac(mac_raw) if mac_raw else None
                sn = (sn_raw or "").strip().upper() or None
                if not mac and not sn:
                    continue
                if sn and sn in seen_sns:
                    continue
                if mac and mac in seen_macs:
                    continue
                if sn:
                    seen_sns.add(sn)
                if mac:
                    seen_macs.add(mac)
                pairs.append({"mac": mac, "sn": sn})
        if pairs:
            # Filtra os que já existem (por SN ou MAC real)
            sns_list = [p_["sn"] for p_ in pairs if p_.get("sn")]
            macs_list = [p_["mac"] for p_ in pairs if p_.get("mac")]
            q_or: List[Dict[str, Any]] = []
            if sns_list:
                q_or.append({"scan_sn": {"$in": sns_list}})
            if macs_list:
                q_or.append({"mac": {"$in": macs_list}})
            existing_keys_sn: set[str] = set()
            existing_keys_mac: set[str] = set()
            if q_or:
                async for ex in db.stok_onts.find(
                    {"company_id": cid, "$or": q_or},
                    {"_id": 0, "mac": 1, "scan_sn": 1},
                ):
                    if ex.get("scan_sn"):
                        existing_keys_sn.add(ex["scan_sn"])
                    if ex.get("mac"):
                        existing_keys_mac.add(ex["mac"])
            new_pairs = [pr for pr in pairs
                         if (pr.get("sn") not in existing_keys_sn or not pr.get("sn"))
                         and (pr.get("mac") not in existing_keys_mac or not pr.get("mac"))]
            model = (p.get("items") or [{}])[0].get("description") or "ONT"
            docs = []
            for pr in new_pairs:
                # MAC placeholder quando só tem SN (segue convenção SN-... iter174)
                final_mac = pr.get("mac") or (f"SN-{pr['sn']}" if pr.get("sn") else None)
                if not final_mac:
                    continue
                docs.append({
                    "company_id": cid, "mac": final_mac, "model": model[:120],
                    "scan_sn": pr.get("sn"),
                    "location_type": "empresa", "location_id": "empresa",
                    "praca_id": p.get("praca_id"),
                    "warehouse_responsible_id":
                        p.get("responsible_collaborator_id"),
                    "purchase_id": purchase_id,
                    "client_name": None, "status": "disponivel",
                    "source": "compra",
                    "created_by": user.get("email", "?"),
                    "created_at": now_iso(),
                })
            if docs:
                await db.stok_onts.insert_many([dict(d) for d in docs])
            items_imported = len(docs)
            macs_imported = [d["mac"] for d in docs]
            skipped = (len(existing_keys_sn) + len(existing_keys_mac))
            if skipped:
                notes.append(
                    f"{skipped} ONT(s) já cadastradas (SN/MAC) — ignoradas")

    elif p["type"] == "insumo":
        # Importa catálogo + helper de histórico do módulo stok
        from routes.stok import (
            CONSUMABLE_CATALOG, CONSUMABLE_BY_ID, _add_history,
        )

        def _match_consumable(desc: str) -> Optional[Dict[str, Any]]:
            """Tenta casar a descrição da compra com um item do catálogo."""
            d = (desc or "").lower()
            # 1) match direto pelo nome do catálogo
            for c in CONSUMABLE_CATALOG:
                if c["name"].lower() in d:
                    return c
            # 2) fallback por palavras-chave
            if "drop" in d:
                return CONSUMABLE_BY_ID.get("drop")
            if "fast" in d:
                return CONSUMABLE_BY_ID.get("conector_fast")
            if "conector" in d and "fibra" in d:
                return CONSUMABLE_BY_ID.get("conector_fibra")
            if "conector" in d and ("rede" in d or "rj45" in d or "rj-45" in d):
                return CONSUMABLE_BY_ID.get("conector_rede")
            if "esticador" in d:
                return CONSUMABLE_BY_ID.get("esticador")
            if "cabo" in d and "rede" in d:
                return CONSUMABLE_BY_ID.get("cabo_rede")
            if "06fo" in d or "6fo" in d or "6 fo" in d:
                return CONSUMABLE_BY_ID.get("fibra_06fo")
            if "12fo" in d or "12 fo" in d:
                return CONSUMABLE_BY_ID.get("fibra_12fo")
            if "24fo" in d or "24 fo" in d:
                return CONSUMABLE_BY_ID.get("fibra_24fo")
            return None

        # Incrementa o estoque DA EMPRESA (location='empresa') usando o
        # schema correto (fields-as-keys), e gera evento `entrada_insumo`
        # em stok_history no formato que o dashboard/balanço entende.
        for it in (p.get("items") or []):
            desc = (it.get("description") or "").strip()
            qty = float(it.get("quantity") or 0)
            if not desc or qty <= 0:
                continue
            match = _match_consumable(desc)
            if not match:
                notes.append(f"Item ignorado (não casou com catálogo): {desc}")
                continue
            cons_id = match["id"]
            cons_name = match["name"]
            cons_unit = match["unit"]
            qty_fmt = int(qty) if float(qty).is_integer() else qty

            await db.stok_stock.update_one(
                {"company_id": cid, "location": "empresa"},
                {"$inc": {cons_id: qty_fmt},
                  "$setOnInsert": {"company_id": cid, "location": "empresa"}},
                upsert=True,
            )
            await _add_history(
                "entrada_insumo",
                (f"Entrada via Central de Compras #{purchase_id[:8]} de "
                 f"{cons_name}: {qty_fmt} {cons_unit}"),
                user.get("name") or user.get("email") or "?",
                "compra",
                cid,
            )
            items_imported += 1

    # else: equipamento/outros = só registra, não afeta stok

    await db.purchases.update_one(
        {"id": purchase_id, "company_id": cid},
        {"$set": {
            "status": "confirmed",
            "confirmed_at": now_iso(),
            "confirmed_by": user.get("email"),
            "items_imported": items_imported,
            "macs_imported": macs_imported,
            "import_notes": notes,
        }},
    )
    return {
        "ok": True,
        "purchase_id": purchase_id,
        "items_imported": items_imported,
        "macs_imported": len(macs_imported),
        "notes": notes,
    }


@router.post("/{purchase_id}/reprocess-from-image")
async def reprocess_from_image(
    purchase_id: str,
    file: UploadFile = File(...),
    item_index: int = 0,
    user: dict = Depends(require_role("gestor")),
) -> Dict[str, Any]:
    """iter211r — Recebe a FOTO da NF novamente e roda a pipeline completa:
       1) Pré-OCR Vision (transcrição literal)
       2) Regex de padrões de SN no texto OCR
       3) Cria as `stok_onts` faltantes

    Útil quando a NF foi importada antes de iter211q e os SNs ficaram
    incompletos. Idempotente: SNs já cadastrados são ignorados.
    """
    cid = user.get("company_id") or DEMO_COMPANY_ID
    p = await db.purchases.find_one(
        {"id": purchase_id, "company_id": cid}, {"_id": 0})
    if not p:
        raise HTTPException(404, "Compra não encontrada.")
    if (p.get("type") or "").lower() != "ont":
        raise HTTPException(400,
            "Reprocessamento por imagem só faz sentido para compras de ONT.")

    raw = await file.read()
    if len(raw) > 15 * 1024 * 1024:
        raise HTTPException(413, "Imagem muito grande (>15MB).")
    fname = (file.filename or "").lower()
    if not any(fname.endswith(e) for e in (".jpg", ".jpeg", ".png", ".webp")):
        raise HTTPException(415, "Envie uma imagem (JPG/PNG/WEBP).")

    image_b64 = base64.b64encode(raw).decode()
    text_content = ""

    # 1) Pré-OCR Vision (transcrição literal)
    try:
        from emergentintegrations.llm.chat import (
            ImageContent, LlmChat, UserMessage,
        )
        from services.ai_keys import resolve_keys
        keys = await resolve_keys(cid)
        ai_key = (keys.get("anthropic") or keys.get("openai")
                       or keys.get("gemini"))
        if not ai_key:
            raise HTTPException(503,
                "IA não configurada. Use o botão 🔄 Reprocessar SNs e cole "
                "os SNs manualmente.")
        ocr_prompt = (
            "Transcreva TODO o texto visível desta nota fiscal/recibo, "
            "linha por linha, preservando a ordem vertical. NÃO interprete "
            "nem traduza. Inclua TODOS os números de série (Nº Série/SN/S/N) "
            "na íntegra, mesmo se forem dezenas. NÃO use markdown. "
            "NÃO use JSON. Apenas o TEXTO BRUTO transcrito.")
        chat = LlmChat(
            api_key=ai_key,
            session_id=f"reproc-ocr-{cid}-{uuid.uuid4().hex[:6]}",
            system_message="Você transcreve texto de imagens fielmente.",
        ).with_model("anthropic", "claude-sonnet-4-5")
        ocr_resp = await chat.send_message(UserMessage(
            text=ocr_prompt,
            file_contents=[ImageContent(image_base64=image_b64)],
        ))
        text_content = (ocr_resp or "").strip()
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("[purchases iter211r] OCR Vision falhou: %s", e)
        raise HTTPException(500, f"OCR falhou: {e!s}")

    # 2) Regex de SNs
    import re as _re
    SN_RE = _re.compile(r"\b[A-Z]{2,6}[A-Z0-9]{6,20}\b")
    STOPWORDS = {"CNPJ", "RUA", "AVENIDA", "CEP", "RECIBO",
                  "ORDEM", "FORNECEDOR", "CLIENTE", "TELEFONE",
                  "ITAU", "GPON", "WIFI", "AC1200", "BIDI"}
    candidates: List[str] = []
    for m in SN_RE.finditer((text_content or "").upper()):
        tok = m.group(0)
        if tok in STOPWORDS or len(tok) <= 8:
            continue
        if tok not in candidates:
            candidates.append(tok)
    if not candidates:
        return {"ok": False, "inserted": 0,
                 "message": "Nenhum SN reconhecido na imagem. "
                              "Tente uma foto mais nítida.",
                 "ocr_chars": len(text_content)}

    # 3) Reaproveita lógica do reprocess-sns (insere candidates como SNs)
    items = p.get("items") or [{}]
    if item_index >= len(items):
        item_index = 0
    item = items[item_index]
    existing = set()
    async for ex in db.stok_onts.find(
        {"company_id": cid, "scan_sn": {"$in": candidates}},
        {"_id": 0, "scan_sn": 1},
    ):
        if ex.get("scan_sn"):
            existing.add(ex["scan_sn"])
    new_sns = [sn for sn in candidates if sn not in existing]
    if not new_sns:
        return {"ok": True, "inserted": 0,
                 "already_existed": len(existing),
                 "candidates_found": len(candidates),
                 "message": (f"{len(candidates)} SN(s) extraídos da foto, "
                              "mas todos já estão no estoque.")}
    model = (item.get("description") or "ONT")[:120]
    docs = []
    for sn in new_sns:
        docs.append({
            "company_id": cid,
            "mac": f"SN-{sn}",
            "scan_sn": sn,
            "model": model,
            "location_type": "empresa", "location_id": "empresa",
            "praca_id": p.get("praca_id"),
            "warehouse_responsible_id": p.get("responsible_collaborator_id"),
            "purchase_id": purchase_id,
            "client_name": None, "status": "disponivel",
            "source": "reprocess-image",
            "created_by": user.get("email", "?"),
            "created_at": now_iso(),
        })
    await db.stok_onts.insert_many([dict(d) for d in docs])
    cur_sns = list(item.get("sns") or [])
    for sn in new_sns:
        if sn not in cur_sns:
            cur_sns.append(sn)
    items[item_index]["sns"] = cur_sns
    await db.purchases.update_one(
        {"id": purchase_id, "company_id": cid},
        {"$set": {"items": items,
                    "raw_text_preview": text_content[:3000],
                    "reprocess_at": now_iso(),
                    "reprocess_by": user.get("email")}},
    )
    return {
        "ok": True,
        "inserted": len(new_sns),
        "already_existed": len(existing),
        "candidates_found": len(candidates),
        "sns": new_sns,
        "message": (f"{len(new_sns)} ONT(s) adicionadas ao estoque "
                      f"a partir da foto."),
    }


class PurchaseReprocessIn(BaseModel):
    """iter211o — entrada do reprocessamento de SNs.

    Se `extra_sns` for fornecido, ANEXA esses SNs ao item ONT da compra
    (útil quando o usuário tem a foto/texto da NF e quer digitar/colar
    manualmente os SNs que faltaram). Se vazio, usa o regex no
    raw_text_preview já armazenado.
    """
    extra_sns: Optional[List[str]] = None
    item_index: Optional[int] = None  # qual item da compra anexar (default 0)
    model: Optional[str] = None        # opcional: força o modelo da ONT


@router.post("/{purchase_id}/reprocess-sns")
async def reprocess_purchase_sns(
    purchase_id: str,
    body: PurchaseReprocessIn = PurchaseReprocessIn(),
    user: dict = Depends(require_role("gestor")),
) -> Dict[str, Any]:
    """iter211o — Reprocessa SNs de uma compra ONT já lançada.

    Caso de uso real (NF Baixadanet #284306 do recibo Fiberhome):
      - A IA inicial perdeu os 10 SNs verticalmente listados.
      - Compra confirmada com 0 ONTs no estoque.
      - Usuário chama este endpoint passando `extra_sns: [10 SNs]`.
      - Sistema cria as 10 stok_onts no estoque da empresa, vinculadas
        ao purchase_id pra rastreabilidade.

    Idempotente: se um SN já existe no `stok_onts`, é ignorado (não duplica).
    """
    cid = user.get("company_id") or DEMO_COMPANY_ID
    p = await db.purchases.find_one(
        {"id": purchase_id, "company_id": cid}, {"_id": 0})
    if not p:
        raise HTTPException(404, "Compra não encontrada.")
    if (p.get("type") or "").lower() != "ont":
        raise HTTPException(400,
            "Reprocessamento de SNs só faz sentido para compras de ONT.")

    items = p.get("items") or [{}]
    idx = body.item_index if body.item_index is not None else 0
    if idx >= len(items):
        raise HTTPException(400, f"item_index={idx} inválido (items={len(items)}).")
    item = items[idx]

    # Colete SNs candidatos:
    candidates: List[str] = []
    # 1) Vindos no body
    for s in (body.extra_sns or []):
        sv = (s or "").strip().upper()
        if sv and sv not in candidates:
            candidates.append(sv)
    # 2) Regex no raw_text_preview se nada veio no body
    if not candidates and p.get("raw_text_preview"):
        import re as _re
        SN_RE = _re.compile(r"\b[A-Z]{2,6}[A-Z0-9]{6,20}\b")
        STOPWORDS = {"CNPJ", "RUA", "AVENIDA", "CEP", "RECIBO",
                      "ORDEM", "FORNECEDOR", "CLIENTE", "TELEFONE",
                      "ITAU", "GPON", "WIFI", "AC1200", "BIDI"}
        for m in SN_RE.finditer(p["raw_text_preview"].upper()):
            tok = m.group(0)
            if tok in STOPWORDS or len(tok) <= 8:
                continue
            if tok not in candidates:
                candidates.append(tok)

    if not candidates:
        raise HTTPException(400,
            "Nenhum SN para processar. Forneça `extra_sns` ou garanta que "
            "a NF tem texto OCR.")

    # Filtra SNs já cadastrados na empresa
    existing = set()
    async for ex in db.stok_onts.find(
        {"company_id": cid, "scan_sn": {"$in": candidates}},
        {"_id": 0, "scan_sn": 1},
    ):
        if ex.get("scan_sn"):
            existing.add(ex["scan_sn"])
    new_sns = [sn for sn in candidates if sn not in existing]

    if not new_sns:
        return {"ok": True, "inserted": 0,
                 "already_existed": len(existing),
                 "message": "Todos os SNs já estão no estoque."}

    model = (body.model or item.get("description") or "ONT")[:120]
    docs = []
    for sn in new_sns:
        docs.append({
            "company_id": cid,
            "mac": f"SN-{sn}",  # placeholder estável até SmartOLT aprovisionar
            "scan_sn": sn,
            "model": model,
            "location_type": "empresa", "location_id": "empresa",
            "praca_id": p.get("praca_id"),
            "warehouse_responsible_id": p.get("responsible_collaborator_id"),
            "purchase_id": purchase_id,
            "client_name": None, "status": "disponivel",
            "source": "compra-reprocess",
            "created_by": user.get("email", "?"),
            "created_at": now_iso(),
        })
    await db.stok_onts.insert_many([dict(d) for d in docs])

    # Atualiza a purchase com os SNs anexados
    cur_sns = list(item.get("sns") or [])
    for sn in new_sns:
        if sn not in cur_sns:
            cur_sns.append(sn)
    items[idx]["sns"] = cur_sns
    await db.purchases.update_one(
        {"id": purchase_id, "company_id": cid},
        {"$set": {"items": items,
                    "reprocess_at": now_iso(),
                    "reprocess_by": user.get("email")}},
    )
    return {
        "ok": True,
        "inserted": len(new_sns),
        "already_existed": len(existing),
        "sns": new_sns,
        "message": f"{len(new_sns)} ONT(s) adicionadas ao estoque da empresa.",
    }


@router.delete("/{purchase_id}")
async def delete_purchase(
    purchase_id: str,
    user: dict = Depends(require_role("auditor")),
) -> Dict[str, Any]:
    """iter177 — Apaga uma compra. Apenas AUDITOR pode executar (acima de
    gestor/administrador). Se a compra estava `confirmed`, REVERTE o
    impacto no estoque:
      - ONTs: apaga `stok_onts` que ainda estão `disponivel` na empresa
        com `purchase_id` correspondente. ONTs já em uso (cliente/técnico)
        permanecem (não destrutivo).
      - Insumos: decrementa `stok_stock.empresa` e grava `entrada_insumo_reversao`
        em `stok_history`.
    Tudo é logado em `purchases_deletion_audit` para rastreabilidade.
    """
    cid = user.get("company_id") or DEMO_COMPANY_ID
    p = await db.purchases.find_one({"id": purchase_id, "company_id": cid},
                                          {"_id": 0})
    if not p:
        raise HTTPException(404, "Compra não encontrada")

    reverted_summary = await _revert_purchase_stock_impact(cid, p, user)

    await db.purchases.delete_one({"id": purchase_id, "company_id": cid})
    await db.purchases_deletion_audit.insert_one({
        "id": f"pda-{uuid.uuid4().hex[:12]}",
        "company_id": cid,
        "deleted_purchase_id": purchase_id,
        "deleted_purchase_snapshot": p,
        "deleted_by_email": user.get("email"),
        "deleted_by_name": user.get("name"),
        "deleted_by_role": user.get("role"),
        "deleted_at": now_iso(),
        "reverted_summary": reverted_summary,
    })
    return {
        "ok": True,
        "purchase_id": purchase_id,
        "reverted": reverted_summary,
    }


class BatchDeleteIn(BaseModel):
    ids: List[str]


@router.post("/batch-delete")
async def batch_delete_purchases(
    payload: BatchDeleteIn,
    user: dict = Depends(require_role("auditor")),
) -> Dict[str, Any]:
    """iter177 — Apaga várias compras em lote (auditor). Retorna resumo
    por id (sucesso/erro). Idêntico a `DELETE /{id}` por id, mas atômico
    em um único request para o frontend.
    """
    cid = user.get("company_id") or DEMO_COMPANY_ID
    results: List[Dict[str, Any]] = []
    for pid in (payload.ids or []):
        try:
            p = await db.purchases.find_one({"id": pid, "company_id": cid},
                                                  {"_id": 0})
            if not p:
                results.append({"id": pid, "ok": False, "error": "not_found"})
                continue
            reverted = await _revert_purchase_stock_impact(cid, p, user)
            await db.purchases.delete_one({"id": pid, "company_id": cid})
            await db.purchases_deletion_audit.insert_one({
                "id": f"pda-{uuid.uuid4().hex[:12]}",
                "company_id": cid,
                "deleted_purchase_id": pid,
                "deleted_purchase_snapshot": p,
                "deleted_by_email": user.get("email"),
                "deleted_by_name": user.get("name"),
                "deleted_by_role": user.get("role"),
                "deleted_at": now_iso(),
                "reverted_summary": reverted,
                "batch": True,
            })
            results.append({"id": pid, "ok": True, "reverted": reverted})
        except Exception as e:
            logger.warning("[batch_delete] %s falhou: %s", pid, e)
            results.append({"id": pid, "ok": False, "error": str(e)[:200]})
    return {
        "processed": len(payload.ids or []),
        "succeeded": sum(1 for r in results if r["ok"]),
        "failed": sum(1 for r in results if not r["ok"]),
        "results": results,
    }


async def _revert_purchase_stock_impact(cid: str, p: Dict[str, Any],
                                            user: dict) -> Dict[str, Any]:
    """Reverte o impacto de uma compra confirmada no estoque. Não destrutivo:
    ONTs já em uso por cliente/técnico PERMANECEM (apenas o vínculo
    `purchase_id` é mantido para histórico, mas a compra é apagada)."""
    summary = {"onts_deleted": 0, "onts_skipped_in_use": 0,
                  "insumos_reverted": []}
    if p.get("status") != "confirmed":
        return {**summary, "skipped": "purchase_not_confirmed"}

    ptype = p.get("type")
    pid = p["id"]
    if ptype == "ont":
        # Apaga ONTs disponíveis com este purchase_id
        in_use = await db.stok_onts.count_documents({
            "company_id": cid, "purchase_id": pid,
            "location_type": {"$ne": "empresa"},
        })
        summary["onts_skipped_in_use"] = in_use
        res = await db.stok_onts.delete_many({
            "company_id": cid, "purchase_id": pid,
            "location_type": "empresa", "status": "disponivel",
        })
        summary["onts_deleted"] = res.deleted_count
    elif ptype == "insumo":
        from routes.stok import (
            CONSUMABLE_CATALOG, CONSUMABLE_BY_ID, _add_history,
        )

        def _match(desc: str):
            d = (desc or "").lower()
            for c in CONSUMABLE_CATALOG:
                if c["name"].lower() in d:
                    return c
            if "drop" in d:
                return CONSUMABLE_BY_ID.get("drop")
            if "fast" in d:
                return CONSUMABLE_BY_ID.get("conector_fast")
            if "conector" in d and "fibra" in d:
                return CONSUMABLE_BY_ID.get("conector_fibra")
            if "conector" in d and ("rede" in d or "rj45" in d):
                return CONSUMABLE_BY_ID.get("conector_rede")
            if "esticador" in d:
                return CONSUMABLE_BY_ID.get("esticador")
            if "cabo" in d and "rede" in d:
                return CONSUMABLE_BY_ID.get("cabo_rede")
            if "06fo" in d or "6fo" in d:
                return CONSUMABLE_BY_ID.get("fibra_06fo")
            if "12fo" in d:
                return CONSUMABLE_BY_ID.get("fibra_12fo")
            if "24fo" in d:
                return CONSUMABLE_BY_ID.get("fibra_24fo")
            return None

        for it in (p.get("items") or []):
            desc = (it.get("description") or "").strip()
            qty = float(it.get("quantity") or 0)
            if not desc or qty <= 0:
                continue
            m = _match(desc)
            if not m:
                continue
            cons_id = m["id"]
            qty_fmt = int(qty) if float(qty).is_integer() else qty
            await db.stok_stock.update_one(
                {"company_id": cid, "location": "empresa"},
                {"$inc": {cons_id: -qty_fmt}},
            )
            await _add_history(
                "entrada_insumo_reversao",
                (f"REVERSÃO — Compra apagada pelo auditor "
                 f"{user.get('name') or user.get('email')} · "
                 f"{m['name']}: -{qty_fmt} {m['unit']}"),
                user.get("name") or user.get("email") or "?",
                "estorno",
                cid,
            )
            summary["insumos_reverted"].append({
                "consumable_id": cons_id, "qty": qty_fmt})
    return summary
